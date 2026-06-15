"""
Confirmed roster allowlist: ingest and DB access only for these athletes.

Source workbook (default via ROSTER_ALLOWLIST_XLSX): sheet "GymAware" with columns
Last Name, First Name, GymAware API ID, Vald tenant_ID, Vald Profile_ID, ...

Enable filtering: ROSTER_FILTER=1 (see .env.example).
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

_PROJECT_ROOT = Path(__file__).resolve().parents[1]

DEFAULT_ROSTER_XLSX = "roster_new.xlsx"
LEGACY_ROSTER_XLSX = "Updated Athelete Reference IDs.xlsx"


def _project_root() -> Path:
    return _PROJECT_ROOT


def env_roster_filter_enabled() -> bool:
    v = os.getenv("ROSTER_FILTER", "").strip().lower()
    return v in ("1", "true", "yes", "on")


def roster_allowlist_path() -> Path:
    p = os.getenv("ROSTER_ALLOWLIST_XLSX", "").strip()
    if p:
        return Path(p)
    # Committed copy for CI / clone-and-run (see data/roster/README.md)
    committed = _project_root() / "data" / "roster" / DEFAULT_ROSTER_XLSX
    if committed.is_file():
        return committed
    # Toolkit root, repo parent, then legacy workbook name
    for name in (DEFAULT_ROSTER_XLSX, LEGACY_ROSTER_XLSX):
        here = _project_root() / name
        if here.is_file():
            return here
        parent = _project_root().parent / name
        if parent.is_file():
            return parent
    return committed


def _norm_cell(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _empty_if_placeholder(x: Any) -> str:
    """Treat N/A, NA, -, TBD, etc. as empty for optional roster cells (jersey, etc.)."""
    s = _norm_cell(x)
    if not s:
        return ""
    low = s.casefold()
    if low in (
        "n/a",
        "n.a.",
        "n.a",
        "none",
        "null",
        "-",
        "--",
        "not available",
        "tbd",
        "tba",
    ) or low == "nan":
        return ""
    return s


def _norm_header(x: Any) -> str:
    return _norm_cell(x).lower()


def _find_col(header: list[str], *needles: str) -> int | None:
    for i, h in enumerate(header):
        if all(n in h for n in needles):
            return i
    return None


def _parse_whoop_user_id(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and str(v) == "nan"):
        return None
    s = _empty_if_placeholder(v)
    if not s:
        return None
    if s.isdigit():
        return s
    m = re.search(r"\d{4,}", s)
    return m.group(0) if m else None


def _parse_uuid_cell(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and str(v) == "nan"):
        return None
    s = _empty_if_placeholder(v)
    if not s:
        return None
    m = re.search(
        r"[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}",
        s,
    )
    if m:
        return m.group(0).lower()
    return None


@dataclass(frozen=True)
class RosterAllowlist:
    """IDs from the client roster workbook + optional explicit Catapult athlete UUIDs."""

    gymaware_refs: frozenset[int]
    vald_profile_ids: frozenset[str]
    catapult_athlete_ids: frozenset[str]
    catapult_jerseys: frozenset[str]


def load_roster_allowlist(path: str | Path | None = None) -> tuple[list[dict[str, Any]], RosterAllowlist]:
    """
    Parse the GymAware sheet. Returns (row dicts for tooling, allowlist sets).

    Row dict keys: last_name, first_name, athlete_reference (int), vald_profile_id (optional str).
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("Install openpyxl: pip install openpyxl") from e

    path = Path(path) if path else roster_allowlist_path()
    if not path.is_file():
        raise FileNotFoundError(f"Roster workbook not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = "GymAware" if "GymAware" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows_iter = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows_iter:
        return [], RosterAllowlist(frozenset(), frozenset(), frozenset(), frozenset())

    header = [_norm_header(c) for c in rows_iter[0]]
    first_row_text = " ".join(header)
    modern = "gymaware" in first_row_text and "api" in first_row_text

    rows_out: list[dict[str, Any]] = []
    grefs: set[int] = set()
    valds: set[str] = set()
    cats: set[str] = set()
    jerseys: set[str] = set()

    data_rows = rows_iter[1:] if modern else rows_iter

    for row in data_rows:
        if not row:
            continue
        if modern:
            i_ln = _find_col(header, "last", "name")
            i_fn = _find_col(header, "first", "name")
            i_ga = _find_col(header, "gymaware", "api")
            if i_ga is None:
                i_ga = _find_col(header, "gymaware")
            i_vald = _find_col(header, "vald", "profile")
            i_cat_uuid = _find_col(header, "catapult", "athlete")
            i_cat_jersey = _find_col(header, "catapult", "jersey")
            i_whoop = _find_col(header, "whoop")
            i_global = _find_col(header, "global", "athlete")
            if i_global is None:
                i_global = _find_col(header, "internal", "key")
            if i_ln is None or i_fn is None or i_ga is None:
                continue
            last = row[i_ln] if i_ln < len(row) else None
            first = row[i_fn] if i_fn < len(row) else None
            ref = row[i_ga] if i_ga < len(row) else None
            vald_cell = row[i_vald] if i_vald is not None and i_vald < len(row) else None
            cat_cell = row[i_cat_uuid] if i_cat_uuid is not None and i_cat_uuid < len(row) else None
            jersey_cell = row[i_cat_jersey] if i_cat_jersey is not None and i_cat_jersey < len(row) else None
            whoop_cell = row[i_whoop] if i_whoop is not None and i_whoop < len(row) else None
            global_cell = row[i_global] if i_global is not None and i_global < len(row) else None
        else:
            if len(row) < 4:
                continue
            last = row[1]
            first = row[2]
            ref = row[3]
            vald_cell = None
            cat_cell = None
            jersey_cell = None
            whoop_cell = None
            global_cell = None

        if ref is None or str(ref).strip() == "":
            continue
        if isinstance(ref, str) and "API ID" in ref:
            continue
        try:
            if isinstance(ref, bool):
                continue
            if isinstance(ref, int):
                ref_int = ref
            elif isinstance(ref, float):
                ref_int = int(ref)
            else:
                ref_int = int(str(ref).strip())
        except (TypeError, ValueError):
            continue

        ln = _norm_cell(last)
        fn = _norm_cell(first)
        rows_out.append(
            {
                "last_name": ln,
                "first_name": fn,
                "athlete_reference": ref_int,
            }
        )
        grefs.add(ref_int)

        vu = _parse_uuid_cell(vald_cell)
        if vu:
            valds.add(vu)
            rows_out[-1]["vald_profile_id"] = vu

        if cat_cell is not None:
            cu = _parse_uuid_cell(cat_cell)
            if cu:
                cats.add(cu)
                rows_out[-1]["catapult_athlete_id"] = cu

        if jersey_cell is not None:
            js = _empty_if_placeholder(jersey_cell)
            if js:
                jerseys.add(js)
                rows_out[-1]["catapult_jersey"] = js

        wu = _parse_whoop_user_id(whoop_cell)
        if wu:
            rows_out[-1]["whoop_user_id"] = wu

        if global_cell is not None:
            gk = _empty_if_placeholder(global_cell)
            if gk:
                rows_out[-1]["internal_key"] = gk

    allow = RosterAllowlist(
        frozenset(grefs),
        frozenset(valds),
        frozenset(cats),
        frozenset(jerseys),
    )
    return rows_out, allow


def load_catapult_athlete_ids_from_identity(
    database_url: str,
    gymaware_refs: set[int],
) -> set[str]:
    """Resolve Catapult UUID strings from public.athlete_identity for roster GymAware IDs."""
    if not database_url.strip() or not gymaware_refs:
        return set()
    import psycopg2

    conn = psycopg2.connect(database_url)
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT btrim(catapult_athlete_id)
            FROM public.athlete_identity
            WHERE gymaware_athlete_reference = ANY(%s)
              AND catapult_athlete_id IS NOT NULL
              AND btrim(catapult_athlete_id) <> ''
            """,
            (list(gymaware_refs),),
        )
        return {r[0] for r in cur.fetchall() if r[0]}
    finally:
        conn.close()


def resolved_catapult_athlete_allowlist(
    database_url: str,
    roster: RosterAllowlist,
) -> set[str]:
    """Explicit Catapult IDs from the workbook plus athlete_identity crosswalk."""
    from_db = load_catapult_athlete_ids_from_identity(database_url, set(roster.gymaware_refs))
    out = set(roster.catapult_athlete_ids) | from_db
    return out


def catapult_roster_filters(
    database_url: str,
    roster: RosterAllowlist,
) -> tuple[set[str] | None, set[str] | None]:
    """
    Returns (allowed_uuids, allowed_jerseys) for Catapult /stats filtering.
    Prefer jersey labels when the workbook lists Catapult Jerseys (teams often have no stable athlete UUID).
    Case-insensitive jersey match uses casefold().
    """
    jersey_set = {j.strip() for j in roster.catapult_jerseys if j and str(j).strip()}
    if jersey_set:
        fold = {j.casefold() for j in jersey_set}
        return (None, fold)

    uuids = resolved_catapult_athlete_allowlist(database_url, roster)
    if uuids:
        low = {u.strip().lower() for u in uuids if u and str(u).strip()}
        return (low, None)

    return (None, None)


def whoop_allowed_state_labels(roster: RosterAllowlist) -> set[str]:
    """state_label in whoop_oauth_token matches GymAware ID string from OAuth start link."""
    return {str(x) for x in roster.gymaware_refs}
